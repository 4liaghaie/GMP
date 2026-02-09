# app/views.py
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
from .filters import HSCodeFilter  # Import the filter set

from django.db import transaction
from django.utils.text import slugify
from rest_framework import viewsets, status, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAdminUser
from .serializers import HSCodeSerializer
from openpyxl import load_workbook
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend

from .models import Season, Heading, HSCode


# ----------------------------
# Helpers
# ----------------------------
class CustomPageNumberPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 600
def _norm_header(h: str) -> str:
    # normalize header names like "Season Code" -> "season_code"
    return slugify(str(h or "")).replace("-", "_").strip().lower()

def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s

def _to_int_or_none(v: Any) -> Optional[int]:
    s = _clean_str(v)
    if s == "":
        return None
    try:
        return int(float(s))  # allows "12.0" from excel
    except ValueError:
        return None

def _read_rows(file_obj) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Returns (headers, rows_as_dicts)
    Supports .csv and .xlsx (first sheet)
    """
    name = (getattr(file_obj, "name", "") or "").lower()

    if name.endswith(".csv"):
        raw = file_obj.read()
        # handle utf-8 with possible BOM
        text = raw.decode("utf-8-sig", errors="replace")
        f = io.StringIO(text)
        reader = csv.reader(f)
        all_rows = list(reader)
        if not all_rows:
            return [], []

        headers = [_norm_header(h) for h in all_rows[0]]
        data_rows = []
        for r in all_rows[1:]:
            if not any(_clean_str(x) for x in r):
                continue
            row = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
            data_rows.append(row)
        return headers, data_rows

    if name.endswith(".xlsx"):
        wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            first = next(rows_iter)
        except StopIteration:
            return [], []

        headers = [_norm_header(h) for h in first]
        data_rows = []
        for r in rows_iter:
            if r is None:
                continue
            if not any(_clean_str(x) for x in r):
                continue
            row = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
            data_rows.append(row)
        return headers, data_rows

    raise ValueError("Unsupported file type. Upload .csv or .xlsx")


@dataclass
class ImportReport:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    row_errors: List[Dict[str, Any]] = None

    def __post_init__(self):
        if self.row_errors is None:
            self.row_errors = []


class BaseImportAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAdminUser]  # change if you want
    model_name: str = "UNKNOWN"
    required_columns: List[str] = []

    def post(self, request, *args, **kwargs):
        """
        Form-data:
          - file: CSV/XLSX
          - dry_run: "true"/"false" (optional, default false)
        """
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "file is required (csv/xlsx)."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = str(request.data.get("dry_run", "false")).lower() in ("1", "true", "yes", "y")

        try:
            headers, rows = _read_rows(upload)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        missing = [c for c in self.required_columns if c not in headers]
        if missing:
            return Response(
                {
                    "detail": "Missing required columns.",
                    "missing": missing,
                    "received": headers,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        report = ImportReport()
        # Validate + upsert inside a transaction; if dry_run we rollback
        with transaction.atomic():
            self.process_rows(rows, report)
            if dry_run:
                transaction.set_rollback(True)

        return Response(
            {
                "model": self.model_name,
                "dry_run": dry_run,
                "total_rows": len(rows),
                "created": report.created,
                "updated": report.updated,
                "skipped": report.skipped,
                "errors": report.errors,
                "row_errors": report.row_errors[:200],  # limit payload
            },
            status=status.HTTP_200_OK if report.errors == 0 else status.HTTP_207_MULTI_STATUS,
        )

    def process_rows(self, rows: List[Dict[str, Any]], report: ImportReport) -> None:
        raise NotImplementedError


# ----------------------------
# Season Import
# ----------------------------

class SeasonImportAPIView(BaseImportAPIView):
    model_name = "Season"
    required_columns = ["code"]

    def process_rows(self, rows: List[Dict[str, Any]], report: ImportReport) -> None:
        for idx, r in enumerate(rows, start=2):  # header is row 1
            code = _clean_str(r.get("code"))
            if not code:
                report.skipped += 1
                continue

            defaults = {
                "description": _clean_str(r.get("description")) or None,
                "season_notes": _clean_str(r.get("season_notes")) or None,
            }

            try:
                obj, created = Season.objects.update_or_create(code=code, defaults=defaults)
                if created:
                    report.created += 1
                else:
                    report.updated += 1
            except Exception as e:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": str(e)})


# ----------------------------
# Heading Import
# ----------------------------

class HeadingImportAPIView(BaseImportAPIView):
    model_name = "Heading"
    required_columns = ["code", "season_code"]

    def process_rows(self, rows: List[Dict[str, Any]], report: ImportReport) -> None:
        # cache seasons by code for speed
        season_map = {s.code: s for s in Season.objects.all().only("id", "code")}

        for idx, r in enumerate(rows, start=2):
            code = _clean_str(r.get("code"))
            season_code = _clean_str(r.get("season_code"))
            if not code or not season_code:
                report.skipped += 1
                continue

            season = season_map.get(season_code)
            if not season:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": f"season_code '{season_code}' not found"})
                continue

            defaults = {
                "season": season,
                "description": _clean_str(r.get("description")) or None,
                "heading_notes": _clean_str(r.get("heading_notes")) or None,
            }

            try:
                obj, created = Heading.objects.update_or_create(code=code, defaults=defaults)
                if created:
                    report.created += 1
                else:
                    report.updated += 1
            except Exception as e:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": str(e)})


# ----------------------------
# HSCode Import
# ----------------------------
def derive_season_code_from_hs(code: str) -> str | None:
    """
    HS 'chapter' is first 2 digits. Your Season.code seems to be '1'..'99' (not '01').
    Example: '01xxxxxx' -> '1'
    """
    if not code or len(code) < 2 or not code[:2].isdigit():
        return None
    return str(int(code[:2]))  # '01' -> 1 -> '1'


def derive_heading_code_from_hs(code: str) -> str | None:
    """
    Heading is first 4 digits.
    Example: '01012100' -> '0101'
    """
    if not code or len(code) < 4 or not code[:4].isdigit():
        return None
    return code[:4]

class HSCodeImportAPIView(BaseImportAPIView):
    model_name = "HSCode"
    required_columns = ["code", "goods_name_fa", "goods_name_en", "profit"]  # removed season_code

    def process_rows(self, rows: List[Dict[str, Any]], report: ImportReport) -> None:
        season_map = {s.code: s for s in Season.objects.all().only("id", "code")}
        heading_map = {h.code: h for h in Heading.objects.all().only("id", "code")}

        allowed_suq = {k for (k, _label) in HSCode.SUQ_OPTIONS}

        for idx, r in enumerate(rows, start=2):
            code = _clean_str(r.get("code"))
            if not code:
                report.skipped += 1
                continue

            # ---- derive season + heading from HS code ----
            derived_season_code = derive_season_code_from_hs(code)
            if not derived_season_code:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": "Invalid HS code for season derivation"})
                continue

            season = season_map.get(derived_season_code)
            if not season:
                report.errors += 1
                report.row_errors.append(
                    {"row": idx, "code": code, "error": f"Derived season_code '{derived_season_code}' not found"}
                )
                continue

            derived_heading_code = derive_heading_code_from_hs(code)
            heading = None
            if derived_heading_code:
                heading = heading_map.get(derived_heading_code)  # ok if None


            # ---- SUQ validation (unchanged) ----
            suq = _clean_str(r.get("suq") or r.get("SUQ"))
            if suq == "":
                suq = None
            if suq is not None and suq not in allowed_suq:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": f"Invalid SUQ '{suq}'. Allowed: {sorted(allowed_suq)}"})
                continue

            defaults = {
                "goods_name_fa": _clean_str(r.get("goods_name_fa")),
                "goods_name_en": _clean_str(r.get("goods_name_en")),
                "profit": _clean_str(r.get("profit")),
                "customs_duty_rate": _to_int_or_none(r.get("customs_duty_rate")),
                "import_duty_rate": _clean_str(r.get("import_duty_rate")) or None,
                "priority": _to_int_or_none(r.get("priority")),
                "season": season,
                "heading": heading,
            }
            if suq is not None:
                defaults["SUQ"] = suq

            missing_required = [k for k in ("goods_name_fa", "goods_name_en", "profit") if not defaults.get(k)]
            if missing_required:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": f"Missing required values: {missing_required}"})
                continue

            try:
                obj, created = HSCode.objects.update_or_create(code=code, defaults=defaults)
                report.created += int(created)
                report.updated += int(not created)
            except Exception as e:
                report.errors += 1
                report.row_errors.append({"row": idx, "code": code, "error": str(e)})

class HSCodeViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for managing HSCode objects.
    Provides list, retrieve, create, update, and delete functionality.
    """
    queryset = HSCode.objects.all().order_by("id")
    serializer_class = HSCodeSerializer
    pagination_class = CustomPageNumberPagination  # Apply pagination here
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = HSCodeFilter
    search_fields = ["code", "goods_name_fa", "goods_name_en","heading__description","season__description"]


